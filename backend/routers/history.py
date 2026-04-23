"""History router for viewing sent messages."""
from datetime import datetime, timedelta
from fastapi import APIRouter, Depends, Query, HTTPException
from fastapi.responses import StreamingResponse
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select, func, and_, delete
from typing import List, Optional
import io
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment
from database import get_db
from models.message_log import MessageLog
from schemas.message import MessageResponse

router = APIRouter(prefix="/history", tags=["history"])


@router.get("", response_model=List[MessageResponse])
async def get_history(
    search: Optional[str] = Query(None, description="Search by recipient name, email or phone"),
    channel: Optional[str] = Query(None, pattern="^(whatsapp|email|sms|both)$"),
    status: Optional[str] = Query(None, pattern="^(sent|failed|pending)$"),
    date_from: Optional[datetime] = Query(None, description="Filter from this date"),
    date_to: Optional[datetime] = Query(None, description="Filter until this date"),
    limit: int = Query(50, ge=1, le=500),
    offset: int = Query(0, ge=0),
    db: AsyncSession = Depends(get_db)
):
    """Get message history with filtering options."""
    # Normalize dates for SQLite comparison (naive Bogota time)
    if date_from and date_from.tzinfo:
        date_from = date_from.replace(tzinfo=None)
    if date_to:
        if date_to.tzinfo:
            date_to = date_to.replace(tzinfo=None)
        # If time is exactly midnight, assume they want the whole end day
        if date_to.hour == 0 and date_to.minute == 0:
            date_to = date_to.replace(hour=23, minute=59, second=59)

    query = select(MessageLog)
    
    conditions = []
    
    if search:
        search_pattern = f"%{search}%"
        conditions.append(
            MessageLog.recipient_name.ilike(search_pattern) |
            MessageLog.recipient_email.ilike(search_pattern) |
            MessageLog.recipient_phone.ilike(search_pattern)
        )
    
    if channel:
        conditions.append(MessageLog.channel == channel)
    
    if status:
        conditions.append(MessageLog.status == status)
    
    if date_from:
        conditions.append(MessageLog.sent_at >= date_from)
    
    if date_to:
        conditions.append(MessageLog.sent_at <= date_to)
    
    if conditions:
        query = query.where(and_(*conditions))
    
    query = query.order_by(MessageLog.sent_at.desc())
    query = query.offset(offset).limit(limit)
    
    result = await db.execute(query)
    logs = result.scalars().all()
    
    return logs


@router.get("/stats")
async def get_stats(
    days: int = Query(30, ge=1, le=365, description="Number of days to include in stats"),
    db: AsyncSession = Depends(get_db)
):
    """Get messaging statistics."""
    from models.message_log import get_colombia_time
    cutoff_date = get_colombia_time() - timedelta(days=days)
    
    # Total messages (Histórico basado en el último ID para ignorar borrados)
    total_result = await db.execute(select(func.max(MessageLog.id)))
    total = total_result.scalar() or 0
    
    # Sent messages (últimos 30 días)
    sent_result = await db.execute(
        select(func.count(MessageLog.id)).where(
            and_(MessageLog.sent_at >= cutoff_date, MessageLog.status == "sent")
        )
    )
    sent = sent_result.scalar() or 0
    
    # Failed messages
    failed_result = await db.execute(
        select(func.count(MessageLog.id)).where(
            and_(MessageLog.sent_at >= cutoff_date, MessageLog.status == "failed")
        )
    )
    failed = failed_result.scalar() or 0
    
    # By channel
    whatsapp_result = await db.execute(
        select(func.count(MessageLog.id)).where(
            and_(MessageLog.sent_at >= cutoff_date, MessageLog.channel.in_(["whatsapp", "both"]))
        )
    )
    whatsapp = whatsapp_result.scalar() or 0
    
    email_result = await db.execute(
        select(func.count(MessageLog.id)).where(
            and_(MessageLog.sent_at >= cutoff_date, MessageLog.channel.in_(["email", "both"]))
        )
    )
    email = email_result.scalar() or 0
    
    sms_result = await db.execute(
        select(func.count(MessageLog.id)).where(
            and_(MessageLog.sent_at >= cutoff_date, MessageLog.channel == "sms")
        )
    )
    sms = sms_result.scalar() or 0
    
    return {
        "period_days": days,
        "total": total,
        "sent": sent,
        "failed": failed,
        "success_rate": round((sent / total * 100) if total > 0 else 0, 2),
        "by_channel": {
            "whatsapp": whatsapp,
            "email": email,
            "sms": sms
        }
    }


@router.get("/count")
async def get_history_count(
    channel: Optional[str] = Query(None, pattern="^(whatsapp|email|sms|both)$"),
    status: Optional[str] = Query(None, pattern="^(sent|failed|pending)$"),
    date_from: Optional[datetime] = Query(None, description="Filter from this date"),
    date_to: Optional[datetime] = Query(None, description="Filter until this date"),
    db: AsyncSession = Depends(get_db)
):
    """Get total count of messages in history."""
    # Normalize dates
    if date_from and date_from.tzinfo:
        date_from = date_from.replace(tzinfo=None)
    if date_to:
        if date_to.tzinfo:
            date_to = date_to.replace(tzinfo=None)
        if date_to.hour == 0 and date_to.minute == 0:
            date_to = date_to.replace(hour=23, minute=59, second=59)

    query = select(func.count(MessageLog.id))
    
    conditions = []
    if channel:
        conditions.append(MessageLog.channel == channel)
    
    if status:
        conditions.append(MessageLog.status == status)
        
    if date_from:
        conditions.append(MessageLog.sent_at >= date_from)
        
    if date_to:
        conditions.append(MessageLog.sent_at <= date_to)
        
    if conditions:
        query = query.where(and_(*conditions))
    
    result = await db.execute(query)
    count = result.scalar() or 0
    
    return {"count": count}


@router.get("/export")
async def export_history(
    search: Optional[str] = Query(None, description="Search by recipient name, email or phone"),
    channel: Optional[str] = Query(None, pattern="^(whatsapp|email|sms|both)$"),
    status: Optional[str] = Query(None, pattern="^(sent|failed|pending)$"),
    date_from: Optional[datetime] = Query(None, description="Filter from this date"),
    date_to: Optional[datetime] = Query(None, description="Filter until this date"),
    db: AsyncSession = Depends(get_db)
):
    """Export message history to Excel format based on filters."""
    # Normalize dates
    if date_from and date_from.tzinfo:
        date_from = date_from.replace(tzinfo=None)
    if date_to:
        if date_to.tzinfo:
            date_to = date_to.replace(tzinfo=None)
        if date_to.hour == 0 and date_to.minute == 0:
            date_to = date_to.replace(hour=23, minute=59, second=59)

    query = select(MessageLog)
    
    conditions = []
    
    if search:
        search_pattern = f"%{search}%"
        conditions.append(
            MessageLog.recipient_name.ilike(search_pattern) |
            MessageLog.recipient_email.ilike(search_pattern) |
            MessageLog.recipient_phone.ilike(search_pattern)
        )
    
    if channel:
        conditions.append(MessageLog.channel == channel)
    
    if status:
        conditions.append(MessageLog.status == status)
    
    if date_from:
        conditions.append(MessageLog.sent_at >= date_from)
    
    if date_to:
        conditions.append(MessageLog.sent_at <= date_to)
    
    if conditions:
        query = query.where(and_(*conditions))
    
    query = query.order_by(MessageLog.sent_at.desc())
    
    result = await db.execute(query)
    logs = result.scalars().all()
    
    # Generate Excel
    wb = Workbook()
    ws = wb.active
    ws.title = "Historial"
    
    headers = ["Fecha", "Usuario", "Destinatario", "Canal", "Estado", "Plantilla", "Fecha y Hora"]
    ws.append(headers)
    
    # Styles for headers
    header_fill = PatternFill(start_color="8B5A9B", end_color="8B5A9B", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    for col_idx in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
        
    # Maps for translation
    status_map = {
        "sent": "Enviado",
        "failed": "Fallido",
        "pending": "Pendiente"
    }

    canals_map = {
        "whatsapp": "WhatsApp",
        "email": "Email",
        "sms": "SMS",
        "both": "Ambos"
    }
    
    for log in logs:
        fecha = log.sent_at.strftime("%d/%m/%Y") if log.sent_at else ""
        fecha_hora = log.sent_at.strftime("%d/%m/%Y %H:%M:%S") if log.sent_at else ""
        
        usuario = log.recipient_name or ""
        
        destinatario = ""
        log_channel = log.channel.lower() if log.channel else ""
        
        if log_channel in ("whatsapp", "sms") and log.recipient_phone:
            destinatario = log.recipient_phone
        elif log_channel == "email" and log.recipient_email:
            destinatario = log.recipient_email
        elif log_channel == "both":
            parts = []
            if log.recipient_email: parts.append(log.recipient_email)
            if log.recipient_phone: parts.append(log.recipient_phone)
            destinatario = " / ".join(parts)
        else:
            destinatario = log.recipient_phone or log.recipient_email or ""
             
        canal = canals_map.get(log_channel, log.channel)
        
        log_status = log.status.lower() if log.status else ""
        estado = status_map.get(log_status, log.status)
        
        plantilla = log.subject
        if not plantilla and log_channel == "whatsapp" and log.message_content and log.message_content.startswith("Template: "):
            plantilla = log.message_content.replace("Template: ", "")
            
        if not plantilla:
            plantilla = "N/A"
        
        ws.append([fecha, usuario, destinatario, canal, estado, plantilla, fecha_hora])
        
    # Column widths
    column_widths = {
        "A": 15, # Fecha
        "B": 30, # Usuario
        "C": 35, # Destinatario
        "D": 15, # Canal
        "E": 15, # Estado
        "F": 45, # Plantilla
        "G": 22  # Fecha y Hora
    }
    for col_letter, width in column_widths.items():
        ws.column_dimensions[col_letter].width = width
        
    stream = io.BytesIO()
    wb.save(stream)
    stream.seek(0)
    
    headers_dict = {
        'Content-Disposition': 'attachment; filename="historial.xlsx"'
    }
    
    return StreamingResponse(
        stream, 
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", 
        headers=headers_dict
    )


@router.delete("")
async def delete_history(
    search: Optional[str] = Query(None, description="Borrar por nombre, email o teléfono"),
    date_from: Optional[datetime] = Query(None, description="Borrar desde esta fecha"),
    date_to: Optional[datetime] = Query(None, description="Borrar hasta esta fecha"),
    channel: Optional[str] = Query(None, pattern="^(whatsapp|email|sms|both)$"),
    status: Optional[str] = Query(None, pattern="^(sent|failed|pending)$"),
    db: AsyncSession = Depends(get_db)
):
    """Delete message history based on filters."""
    # Require at least one filter to avoid accidental total wipe
    if not any([search, date_from, date_to, channel, status]):
        raise HTTPException(
            status_code=400, 
            detail="Debe proporcionar al menos un filtro (búsqueda, fecha, canal o estado) para realizar la limpieza"
        )

    # Normalize dates
    if date_from and date_from.tzinfo:
        date_from = date_from.replace(tzinfo=None)
    if date_to:
        if date_to.tzinfo:
            date_to = date_to.replace(tzinfo=None)
        if date_to.hour == 0 and date_to.minute == 0:
            date_to = date_to.replace(hour=23, minute=59, second=59)

    query = delete(MessageLog)
    
    conditions = []
    if search:
        search_pattern = f"%{search}%"
        conditions.append(
            MessageLog.recipient_name.ilike(search_pattern) |
            MessageLog.recipient_email.ilike(search_pattern) |
            MessageLog.recipient_phone.ilike(search_pattern)
        )
        
    if date_from:
        conditions.append(MessageLog.sent_at >= date_from)
    if date_to:
        conditions.append(MessageLog.sent_at <= date_to)
    if channel:
        conditions.append(MessageLog.channel == channel)
    if status:
        conditions.append(MessageLog.status == status)
        
    if conditions:
        query = query.where(and_(*conditions))
    
    result = await db.execute(query)
    await db.commit()
    
    return {
        "message": "Historial limpiado correctamente", 
        "deleted_count": result.rowcount
    }
