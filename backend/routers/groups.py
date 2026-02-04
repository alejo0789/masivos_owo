"""Groups router for managing contact groups."""
import io
import re
from typing import List, Optional
from fastapi import APIRouter, Depends, HTTPException, UploadFile, File, Form
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select, func
from sqlalchemy.orm import selectinload

from database import get_db
from models.group import Group, GroupContact
from schemas.group import (
    GroupCreate, GroupUpdate, GroupResponse, GroupDetailResponse,
    GroupContactCreate, GroupContactUpdate, GroupContactResponse,
    ExcelUploadResponse
)

router = APIRouter(prefix="/groups", tags=["groups"])


def normalize_phone(phone: str) -> str:
    """Normalize phone number to +57 format for Colombia."""
    if not phone:
        return ""
    
    # Remove all non-digit characters
    digits = re.sub(r'\D', '', str(phone))
    
    if not digits:
        return ""
    
    # Handle Colombian numbers
    if digits.startswith('57') and len(digits) >= 12:
        return f"+{digits}"
    elif len(digits) == 10 and digits.startswith('3'):
        # Colombian mobile: 10 digits starting with 3
        return f"+57{digits}"
    elif len(digits) == 7:
        # Colombian landline (7 digits) - add country code
        return f"+57{digits}"
    else:
        # Return as-is with + prefix if not already formatted
        return f"+{digits}" if not phone.startswith('+') else phone


@router.get("", response_model=List[GroupResponse])
async def get_groups(db: AsyncSession = Depends(get_db)):
    """Get all groups with contact counts."""
    # Query groups with contact count
    stmt = (
        select(Group, func.count(GroupContact.id).label("contact_count"))
        .outerjoin(GroupContact)
        .group_by(Group.id)
        .order_by(Group.name)
    )
    result = await db.execute(stmt)
    rows = result.all()
    
    groups = []
    for group, count in rows:
        group_dict = {
            "id": group.id,
            "name": group.name,
            "description": group.description,
            "created_at": group.created_at,
            "updated_at": group.updated_at,
            "contact_count": count
        }
        groups.append(GroupResponse(**group_dict))
    
    return groups


@router.get("/{group_id}", response_model=GroupDetailResponse)
async def get_group(group_id: int, db: AsyncSession = Depends(get_db)):
    """Get a group with all its contacts."""
    stmt = (
        select(Group)
        .options(selectinload(Group.contacts))
        .where(Group.id == group_id)
    )
    result = await db.execute(stmt)
    group = result.scalar_one_or_none()
    
    if not group:
        raise HTTPException(status_code=404, detail="Grupo no encontrado")
    
    return group


@router.post("", response_model=GroupResponse)
async def create_group(
    group_data: GroupCreate,
    db: AsyncSession = Depends(get_db)
):
    """Create a new empty group."""
    # Check if group name already exists
    stmt = select(Group).where(Group.name == group_data.name)
    result = await db.execute(stmt)
    existing = result.scalar_one_or_none()
    
    if existing:
        raise HTTPException(status_code=400, detail="Ya existe un grupo con ese nombre")
    
    group = Group(
        name=group_data.name,
        description=group_data.description
    )
    db.add(group)
    await db.flush()
    
    return GroupResponse(
        id=group.id,
        name=group.name,
        description=group.description,
        created_at=group.created_at,
        updated_at=group.updated_at,
        contact_count=0
    )


@router.post("/upload", response_model=ExcelUploadResponse)
async def upload_excel(
    file: UploadFile = File(...),
    group_name: str = Form(...),
    db: AsyncSession = Depends(get_db)
):
    """
    Upload an Excel file to create a new group with contacts.
    
    The Excel file should have columns: nombre, telefono, correo
    """
    # Validate file type
    if not file.filename.endswith(('.xlsx', '.xls')):
        raise HTTPException(
            status_code=400, 
            detail="El archivo debe ser un Excel (.xlsx o .xls)"
        )
    
    # Check if group name already exists
    stmt = select(Group).where(Group.name == group_name)
    result = await db.execute(stmt)
    existing = result.scalar_one_or_none()
    
    if existing:
        raise HTTPException(status_code=400, detail="Ya existe un grupo con ese nombre")
    
    try:
        # Read the Excel file
        import openpyxl
        
        contents = await file.read()
        wb = openpyxl.load_workbook(io.BytesIO(contents))
        ws = wb.active
        
        # Get headers from first row
        headers = []
        for cell in ws[1]:
            if cell.value:
                headers.append(str(cell.value).lower().strip())
        
        # Map column indices
        col_map = {}
        for i, header in enumerate(headers):
            if header in ['nombre', 'name']:
                col_map['name'] = i
            elif header in ['telefono', 'teléfono', 'phone', 'celular']:
                col_map['phone'] = i
            elif header in ['correo', 'email', 'mail', 'e-mail']:
                col_map['email'] = i
        
        if 'name' not in col_map:
            raise HTTPException(
                status_code=400, 
                detail="El archivo debe tener una columna 'nombre'"
            )
        
        # Create the group
        group = Group(name=group_name)
        db.add(group)
        await db.flush()
        
        # Process rows
        contacts_created = 0
        errors = []
        
        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            try:
                # Get name (required)
                name_idx = col_map.get('name', 0)
                name = row[name_idx] if len(row) > name_idx else None
                
                if not name or str(name).strip() == '':
                    continue  # Skip empty rows
                
                # Get optional fields
                phone_idx = col_map.get('phone')
                phone = str(row[phone_idx]).strip() if phone_idx is not None and len(row) > phone_idx and row[phone_idx] else None
                phone = normalize_phone(phone) if phone else None
                
                email_idx = col_map.get('email')
                email = str(row[email_idx]).strip() if email_idx is not None and len(row) > email_idx and row[email_idx] else None
                
                # Create contact
                contact = GroupContact(
                    group_id=group.id,
                    name=str(name).strip(),
                    phone=phone,
                    email=email
                )
                db.add(contact)
                contacts_created += 1
                
            except Exception as e:
                errors.append(f"Fila {row_idx}: {str(e)}")
        
        await db.flush()
        
        return ExcelUploadResponse(
            group_id=group.id,
            group_name=group.name,
            contacts_created=contacts_created,
            errors=errors
        )
        
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(
            status_code=500,
            detail=f"Error procesando el archivo: {str(e)}"
        )


@router.put("/{group_id}", response_model=GroupResponse)
async def update_group(
    group_id: int,
    group_data: GroupUpdate,
    db: AsyncSession = Depends(get_db)
):
    """Update a group's name or description."""
    stmt = select(Group).where(Group.id == group_id)
    result = await db.execute(stmt)
    group = result.scalar_one_or_none()
    
    if not group:
        raise HTTPException(status_code=404, detail="Grupo no encontrado")
    
    # Check if new name already exists (if changing name)
    if group_data.name and group_data.name != group.name:
        stmt = select(Group).where(Group.name == group_data.name)
        result = await db.execute(stmt)
        existing = result.scalar_one_or_none()
        if existing:
            raise HTTPException(status_code=400, detail="Ya existe un grupo con ese nombre")
        group.name = group_data.name
    
    if group_data.description is not None:
        group.description = group_data.description
    
    await db.flush()
    
    # Get contact count
    stmt = select(func.count(GroupContact.id)).where(GroupContact.group_id == group_id)
    result = await db.execute(stmt)
    count = result.scalar() or 0
    
    return GroupResponse(
        id=group.id,
        name=group.name,
        description=group.description,
        created_at=group.created_at,
        updated_at=group.updated_at,
        contact_count=count
    )


@router.delete("/{group_id}")
async def delete_group(group_id: int, db: AsyncSession = Depends(get_db)):
    """Delete a group and all its contacts."""
    stmt = select(Group).where(Group.id == group_id)
    result = await db.execute(stmt)
    group = result.scalar_one_or_none()
    
    if not group:
        raise HTTPException(status_code=404, detail="Grupo no encontrado")
    
    await db.delete(group)
    await db.flush()
    
    return {"message": "Grupo eliminado correctamente"}


# Contact management endpoints

@router.post("/{group_id}/contacts", response_model=GroupContactResponse)
async def add_contact(
    group_id: int,
    contact_data: GroupContactCreate,
    db: AsyncSession = Depends(get_db)
):
    """Add a contact to a group."""
    # Verify group exists
    stmt = select(Group).where(Group.id == group_id)
    result = await db.execute(stmt)
    group = result.scalar_one_or_none()
    
    if not group:
        raise HTTPException(status_code=404, detail="Grupo no encontrado")
    
    phone = normalize_phone(contact_data.phone) if contact_data.phone else None
    
    contact = GroupContact(
        group_id=group_id,
        name=contact_data.name,
        phone=phone,
        email=contact_data.email
    )
    db.add(contact)
    await db.flush()
    
    return contact


@router.put("/{group_id}/contacts/{contact_id}", response_model=GroupContactResponse)
async def update_contact(
    group_id: int,
    contact_id: int,
    contact_data: GroupContactUpdate,
    db: AsyncSession = Depends(get_db)
):
    """Update a contact in a group."""
    stmt = select(GroupContact).where(
        GroupContact.id == contact_id,
        GroupContact.group_id == group_id
    )
    result = await db.execute(stmt)
    contact = result.scalar_one_or_none()
    
    if not contact:
        raise HTTPException(status_code=404, detail="Contacto no encontrado")
    
    if contact_data.name is not None:
        contact.name = contact_data.name
    if contact_data.phone is not None:
        contact.phone = normalize_phone(contact_data.phone) if contact_data.phone else None
    if contact_data.email is not None:
        contact.email = contact_data.email
    
    await db.flush()
    
    return contact


@router.delete("/{group_id}/contacts/{contact_id}")
async def delete_contact(
    group_id: int,
    contact_id: int,
    db: AsyncSession = Depends(get_db)
):
    """Delete a contact from a group."""
    stmt = select(GroupContact).where(
        GroupContact.id == contact_id,
        GroupContact.group_id == group_id
    )
    result = await db.execute(stmt)
    contact = result.scalar_one_or_none()
    
    if not contact:
        raise HTTPException(status_code=404, detail="Contacto no encontrado")
    
    await db.delete(contact)
    await db.flush()
    
    return {"message": "Contacto eliminado correctamente"}
